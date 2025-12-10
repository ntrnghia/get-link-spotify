"""Excel export for ZingMP3-Spotify results."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXCEL_COLUMN_WIDTHS, EXCEL_HEADERS

# Column mapping: (dict_key, format_fn or None)
COLUMNS = [
    ("rank", None), ("song_name", None), ("artists", None), ("duration", None), ("zing_url", None),
    ("spotify_name", None), ("spotify_artist", None), ("spotify_album", None),
    ("spotify_duration", None), ("spotify_url", None), ("popularity", None),
    ("match_score", lambda v: f"{v * 100:.0f}%" if v else ""),
]


def write_excel(results: list[dict], output_file: str) -> None:
    """Write results to formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ZingMP3 Chart"

    # Header styling
    fills = [
        PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),  # Blue (ZingMP3)
        PatternFill(start_color="1DB954", end_color="1DB954", fill_type="solid"),  # Green (Spotify)
        PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # Orange (Match)
    ]
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font, cell.alignment = header_font, Alignment(horizontal="center")
        cell.fill = fills[0 if col <= 5 else (1 if col <= 11 else 2)]

    # Data rows
    for row_idx, r in enumerate(results, 2):
        for col, (key, fmt) in enumerate(COLUMNS, 1):
            value = r.get(key, "" if key != "popularity" else 0)
            ws.cell(row=row_idx, column=col, value=fmt(value) if fmt else value)

    # Column widths & freeze
    for col, width in enumerate(EXCEL_COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    wb.save(output_file)
