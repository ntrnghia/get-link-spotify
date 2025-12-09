"""
ZingMP3 Chart Crawler with Spotify Links
Parses top 100 songs from chart.html or fetches live from zingmp3.vn using Vietnam proxies
"""

import argparse
import json
import hashlib
import hmac
import os
import re
import time
import warnings
from difflib import SequenceMatcher
from pathlib import Path
from bs4 import BeautifulSoup
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials, SpotifyOAuth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Suppress SSL warnings for proxies
warnings.filterwarnings('ignore', category=InsecureRequestWarning)

# Spotify credentials (from env vars or .env file)
# For local development, create a .env file with:
#   SPOTIFY_CLIENT_ID=your_client_id
#   SPOTIFY_CLIENT_SECRET=your_client_secret
#   SPOTIFY_REFRESH_TOKEN=your_refresh_token (optional, for headless mode)
def _load_env_file():
    """Load environment variables from .env file if it exists."""
    env_file = Path(__file__).parent / '.env'
    if env_file.exists():
        with open(env_file) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ.setdefault(key.strip(), value.strip())

_load_env_file()

SPOTIFY_CLIENT_ID = os.environ.get("SPOTIFY_CLIENT_ID", "")
SPOTIFY_CLIENT_SECRET = os.environ.get("SPOTIFY_CLIENT_SECRET", "")
SPOTIFY_REDIRECT_URI = "http://127.0.0.1:8888/callback"

# ZingMP3 Chart URL
ZINGCHART_URL = "https://zingmp3.vn/zing-chart"

# Proxy sources
GEONODE_API_URL = "https://proxylist.geonode.com/api/proxy-list?country=VN&limit=50&page=1&sort_by=speed&sort_type=asc"
FREE_PROXY_API_URL = "https://api.proxyscrape.com/v4/free-proxy-list/get?request=display_proxies&country=vn&proxy_format=protocolipport&format=json"

# Request headers to mimic browser
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
}

# ZingMP3 API credentials (for weekly charts that require API access)
ZINGMP3_API_KEY = 'X5BM3w8N7MKozC0B85o4KMlzLZKhV00y'
ZINGMP3_SECRET_KEY = 'acOrvUS15XRW2o9JksiK1KgQ6Vbds8ZW'
ZINGMP3_API_VERSION = '1.17.3'


def is_weekly_chart_url(url: str) -> bool:
    """Check if URL is a weekly chart (requires API access)."""
    return 'zing-chart-tuan' in url


def extract_chart_id(url: str) -> str | None:
    """Extract chart ID from ZingMP3 weekly chart URL.

    Example: https://zingmp3.vn/zing-chart-tuan/bai-hat-Viet-Nam/IWZ9Z08I.html -> IWZ9Z08I
    """
    match = re.search(r'/([A-Z0-9]+)\.html$', url)
    return match.group(1) if match else None


def generate_zingmp3_signature(path: str, chart_id: str, ctime: str) -> str:
    """Generate ZingMP3 API signature using HMAC-SHA512."""
    hash_input = f'ctime={ctime}id={chart_id}version={ZINGMP3_API_VERSION}'
    sha256_hash = hashlib.sha256(hash_input.encode()).hexdigest()
    sig = hmac.new(
        ZINGMP3_SECRET_KEY.encode(),
        f'{path}{sha256_hash}'.encode(),
        hashlib.sha512
    ).hexdigest()
    return sig


def fetch_weekly_chart_api(chart_id: str, proxy: dict) -> list[dict] | None:
    """Fetch weekly chart using ZingMP3 API with proxy.

    Returns list of songs or None if failed.
    """
    path = '/api/v2/page/get/week-chart'
    ctime = str(int(time.time()))
    sig = generate_zingmp3_signature(path, chart_id, ctime)

    url = (f'https://zingmp3.vn{path}?id={chart_id}&week=0&year=0'
           f'&ctime={ctime}&version={ZINGMP3_API_VERSION}'
           f'&sig={sig}&apiKey={ZINGMP3_API_KEY}')

    proxy_url = f"http://{proxy['ip']}:{proxy['port']}"
    proxies = {'http': proxy_url, 'https': proxy_url}
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'application/json',
        'Referer': 'https://zingmp3.vn/',
    }

    try:
        response = requests.get(url, headers=headers, proxies=proxies,
                               timeout=20, verify=False)
        data = response.json()

        if data.get('err') == 0:
            return data.get('data', {}).get('items', [])
    except Exception:
        pass

    return None


def parse_api_chart_items(items: list[dict]) -> list[dict]:
    """Parse chart items from ZingMP3 API response.

    Returns list of songs with: position, name, duration, artists
    """
    songs = []
    for i, item in enumerate(items, 1):
        # Parse duration from seconds
        duration_sec = item.get('duration', 0)
        if duration_sec:
            minutes = duration_sec // 60
            seconds = duration_sec % 60
            duration = f"{minutes}:{seconds:02d}"
        else:
            duration = ""

        # Get artists
        artists_names = item.get('artistsNames', '') or ''
        artist_list = [a.strip() for a in artists_names.split(',') if a.strip()]

        songs.append({
            'position': i,
            'name': item.get('title', ''),
            'duration': duration,
            'artists': artists_names,
            'artist_list': artist_list
        })

    return songs


def fetch_vietnam_proxies() -> list[dict]:
    """Fetch Vietnam proxies from multiple sources, sorted by speed (fastest first)."""
    proxies = []

    # Source 1: ProxyScrape API (primary)
    print("  Fetching from ProxyScrape API...")
    try:
        response = requests.get(FREE_PROXY_API_URL, headers=HEADERS, timeout=15)
        if response.status_code == 200:
            data = response.json()
            proxy_list = data.get('proxies', [])
            for proxy in proxy_list:
                proxies.append({
                    'ip': proxy.get('ip', ''),
                    'port': str(proxy.get('port', '')),
                    'speed': proxy.get('timeout', 9999),
                    'source': 'proxyscrape'
                })
            print(f"    Found {len(proxy_list)} proxies from ProxyScrape")
    except Exception as e:
        print(f"    ProxyScrape failed: {e}")

    # Source 3: Free Proxy List (backup - scraping)
    if len(proxies) < 5:
        print("  Fetching from free-proxy-list.net...")
        try:
            response = requests.get(
                "https://free-proxy-list.net/",
                headers=HEADERS,
                timeout=15
            )
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                table = soup.find('table')
                if table:
                    rows = table.find_all('tr')[1:]  # Skip header
                    vn_count = 0
                    for row in rows:
                        cols = row.find_all('td')
                        if len(cols) >= 8:
                            country = cols[2].get_text(strip=True)
                            if country == 'VN' or country == 'Vietnam':
                                proxies.append({
                                    'ip': cols[0].get_text(strip=True),
                                    'port': cols[1].get_text(strip=True),
                                    'speed': 5000,  # Default speed
                                    'source': 'free-proxy-list'
                                })
                                vn_count += 1
                    print(f"    Found {vn_count} Vietnam proxies from free-proxy-list")
        except Exception as e:
            print(f"    free-proxy-list failed: {e}")

    # Remove duplicates and invalid entries
    seen = set()
    unique_proxies = []
    for p in proxies:
        key = f"{p['ip']}:{p['port']}"
        if key not in seen and p['ip'] and p['port']:
            seen.add(key)
            unique_proxies.append(p)

    # Sort by speed (fastest first)
    unique_proxies.sort(key=lambda x: x.get('speed', 9999))

    print(f"  Total unique Vietnam proxies: {len(unique_proxies)}")
    return unique_proxies


def fetch_with_proxy(url: str, proxy: dict, timeout: int = 15) -> str | None:
    """Try to fetch URL using given proxy. Returns HTML content or None."""
    proxy_url = f"http://{proxy['ip']}:{proxy['port']}"
    proxies = {
        'http': proxy_url,
        'https': proxy_url
    }

    try:
        response = requests.get(
            url,
            headers=HEADERS,
            proxies=proxies,
            timeout=timeout,
            verify=False  # Some proxies have SSL issues
        )
        if response.status_code == 200:
            return response.text
    except requests.exceptions.RequestException:
        pass

    return None


def fetch_zingchart_direct(chart_url: str = None) -> str | None:
    """Fetch zing-chart page directly (requires VPN to be connected). Returns HTML content or None."""
    url = chart_url or ZINGCHART_URL
    print(f"\n[VPN MODE] Fetching {url} directly (VPN must be connected)...")

    try:
        response = requests.get(
            url,
            headers=HEADERS,
            timeout=30
        )
        if response.status_code == 200 and 'MusicPlaylist' in response.text:
            print("  SUCCESS! Got chart data.")
            return response.text
        elif response.status_code == 200:
            print("  Page loaded but no chart data (geo-blocked?)")
            return None
        else:
            print(f"  Failed: HTTP {response.status_code}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"  Connection error: {e}")
        return None


def fetch_zingchart_live(chart_url: str = None) -> str | None:
    """Fetch zing-chart page using Vietnam proxies. Returns HTML content or None."""
    url = chart_url or ZINGCHART_URL
    print(f"\n[PROXY MODE] Fetching {url} using Vietnam proxies...")

    proxies = fetch_vietnam_proxies()
    if not proxies:
        print("  No proxies available!")
        return None

    max_tries = min(30, len(proxies))  # Try up to 30 proxies

    # Try each proxy
    for i, proxy in enumerate(proxies[:max_tries], 1):
        source = proxy.get('source', 'unknown')
        speed = proxy.get('speed', '?')
        print(f"  [{i:2d}/{max_tries}] {proxy['ip']}:{proxy['port']} ({source}, {speed}ms)...", end=" ", flush=True)

        html = fetch_with_proxy(url, proxy)

        if html and 'MusicPlaylist' in html:
            print("SUCCESS!")
            return html
        elif html:
            print("No chart data (geo-blocked?)")
        else:
            print("Connection failed")

    print("  All proxies failed!")
    return None


def get_mobile_weekly_url(chart_url: str) -> str:
    """Convert desktop weekly chart URL to mobile URL.

    Example: https://zingmp3.vn/zing-chart-tuan/bai-hat-Viet-Nam/IWZ9Z08I.html
          -> https://m.zingmp3.vn/zing-chart-tuan/bai-hat-Viet-Nam/IWZ9Z08I.html
    """
    return chart_url.replace('://zingmp3.vn/', '://m.zingmp3.vn/')


def parse_mobile_weekly_chart(html: str) -> list[dict] | None:
    """Parse weekly chart from mobile site HTML.

    Returns list of songs with: position, name, duration, artists, artist_list
    """
    soup = BeautifulSoup(html, 'html.parser')

    # Find all chart items
    chart_items = soup.find_all('li', class_='z-chart-item')
    if not chart_items:
        return None

    songs = []
    for item in chart_items:
        # Get rank
        rank_elem = item.find(class_='sort-number')
        rank = rank_elem.get_text(strip=True) if rank_elem else str(len(songs) + 1)

        # Get card-info
        card_info = item.find(class_='card-info')
        if not card_info:
            continue

        # Get title
        title_elem = card_info.find(class_='title')
        title = title_elem.get_text(strip=True) if title_elem else ''

        # Get artist - it's in class='artist'
        artist_elem = card_info.find(class_='artist')
        artist = artist_elem.get_text(strip=True) if artist_elem else ''

        if title:
            # Parse artists into list
            artist_list = [a.strip() for a in artist.split(',') if a.strip()]

            songs.append({
                'position': int(rank) if rank.isdigit() else len(songs) + 1,
                'name': title,
                'duration': '',  # Mobile site doesn't show duration
                'artists': artist,
                'artist_list': artist_list,
                'url': ''  # Mobile site doesn't have URL in this format
            })

    return songs if songs else None


def fetch_weekly_chart_live(chart_url: str) -> list[dict] | None:
    """Fetch weekly chart from mobile site with Vietnam proxies.

    Weekly charts don't have JSON-LD in HTML, so we use the mobile site
    which renders chart data server-side.
    Returns list of songs directly or None if failed.
    """
    mobile_url = get_mobile_weekly_url(chart_url)
    print(f"\n[MOBILE MODE] Fetching weekly chart from {mobile_url}...")

    proxies = fetch_vietnam_proxies()
    if not proxies:
        print("  No proxies available!")
        return None

    max_tries = min(30, len(proxies))  # Try up to 30 proxies

    # Try each proxy
    for i, proxy in enumerate(proxies[:max_tries], 1):
        source = proxy.get('source', 'unknown')
        speed = proxy.get('speed', '?')
        print(f"  [{i:2d}/{max_tries}] {proxy['ip']}:{proxy['port']} ({source}, {speed}ms)...", end=" ", flush=True)

        html = fetch_with_proxy(mobile_url, proxy, timeout=20)

        if html and 'z-chart-item' in html:
            songs = parse_mobile_weekly_chart(html)
            if songs:
                print(f"SUCCESS! Got {len(songs)} songs")
                return songs
            else:
                print("Failed to parse")
        else:
            print("No chart data")

    print("  All proxies failed!")
    return None


def parse_duration(iso_duration: str) -> str:
    """Convert ISO 8601 duration to MM:SS format.

    Example: PT5M20S -> 5:20
    """
    if not iso_duration:
        return ""

    match = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', iso_duration)
    if not match:
        return iso_duration

    hours = int(match.group(1) or 0)
    minutes = int(match.group(2) or 0)
    seconds = int(match.group(3) or 0)

    if hours > 0:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


def parse_chart_content(html_content: str) -> list[dict]:
    """Parse HTML content and extract song data from JSON-LD.

    Returns list of songs with: position, name, duration, artists
    """
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all JSON-LD scripts
    json_ld_scripts = soup.find_all('script', type='application/ld+json')

    songs = []
    for script in json_ld_scripts:
        try:
            data = json.loads(script.string)

            # Look for MusicPlaylist type
            if data.get('@type') == 'MusicPlaylist':
                items = data.get('track', {}).get('itemListElement', [])

                for item in items:
                    song_data = item.get('item', {})
                    artists = song_data.get('byArtist', [])
                    artist_names = [a.get('name', '') for a in artists]

                    songs.append({
                        'position': item.get('position'),
                        'name': song_data.get('name', ''),
                        'duration': parse_duration(song_data.get('duration', '')),
                        'artists': ', '.join(artist_names),
                        'artist_list': artist_names,  # Keep list for search
                        'url': item.get('url', '')  # URL is in itemListElement, not in item
                    })

                break
        except json.JSONDecodeError:
            continue

    return songs


def parse_chart_file(filepath: str) -> list[dict]:
    """Parse chart.html file and extract song data."""
    with open(filepath, 'r', encoding='utf-8') as f:
        html_content = f.read()
    return parse_chart_content(html_content)


def get_spotify_client() -> spotipy.Spotify:
    """Initialize Spotify client with client credentials (read-only)."""
    auth_manager = SpotifyClientCredentials(
        client_id=SPOTIFY_CLIENT_ID,
        client_secret=SPOTIFY_CLIENT_SECRET
    )
    return spotipy.Spotify(auth_manager=auth_manager)


def get_spotify_client_with_auth() -> spotipy.Spotify:
    """Initialize Spotify client with user authorization (for playlist management).

    First time: Opens browser for user login.
    Subsequent runs: Uses cached token from .cache file.
    """
    auth_manager = SpotifyOAuth(
        client_id=SPOTIFY_CLIENT_ID,
        client_secret=SPOTIFY_CLIENT_SECRET,
        redirect_uri=SPOTIFY_REDIRECT_URI,
        scope="playlist-modify-public playlist-modify-private"
    )
    return spotipy.Spotify(auth_manager=auth_manager)


def get_spotify_client_headless() -> spotipy.Spotify:
    """Initialize Spotify client using refresh token (for CI/headless mode).

    Requires SPOTIFY_REFRESH_TOKEN environment variable.
    """
    refresh_token = os.environ.get("SPOTIFY_REFRESH_TOKEN")
    if not refresh_token:
        raise ValueError("SPOTIFY_REFRESH_TOKEN environment variable not set")

    auth_manager = SpotifyOAuth(
        client_id=SPOTIFY_CLIENT_ID,
        client_secret=SPOTIFY_CLIENT_SECRET,
        redirect_uri=SPOTIFY_REDIRECT_URI,
        scope="playlist-modify-public playlist-modify-private",
        open_browser=False
    )

    # Create token info and refresh it
    token_info = auth_manager.refresh_access_token(refresh_token)
    return spotipy.Spotify(auth=token_info['access_token'])


def create_or_get_playlist(sp: spotipy.Spotify, playlist_name: str, chart_url: str = "") -> tuple[str, bool]:
    """Create playlist or get existing one.

    Returns tuple of (playlist_id, is_new).
    """
    user_id = sp.current_user()['id']

    # Search existing playlists
    playlists = sp.current_user_playlists(limit=50)
    while playlists:
        for playlist in playlists['items']:
            if playlist['name'] == playlist_name:
                return playlist['id'], False

        # Get next page
        if playlists['next']:
            playlists = sp.next(playlists)
        else:
            break

    # Create new playlist
    description = f"ZingMP3 Chart - Auto-updated by ZingChart Crawler - {chart_url}" if chart_url else "ZingMP3 Chart - Auto-updated by ZingChart Crawler"
    new_playlist = sp.user_playlist_create(
        user=user_id,
        name=playlist_name,
        public=True,
        description=description
    )
    return new_playlist['id'], True


def update_playlist(sp: spotipy.Spotify, playlist_id: str, track_ids: list[str]) -> int:
    """Replace all tracks in playlist with new ones.

    Returns number of tracks added.
    """
    # Filter out None/empty track IDs
    valid_track_ids = [tid for tid in track_ids if tid]

    if not valid_track_ids:
        return 0

    # Clear existing tracks by replacing with new ones
    # Spotify API allows up to 100 tracks per request
    sp.playlist_replace_items(playlist_id, valid_track_ids[:100])

    # If more than 100 tracks, add the rest
    if len(valid_track_ids) > 100:
        for i in range(100, len(valid_track_ids), 100):
            batch = valid_track_ids[i:i+100]
            sp.playlist_add_items(playlist_id, batch)

    return len(valid_track_ids)


def string_similarity(a: str, b: str) -> float:
    """Calculate string similarity ratio (0.0 to 1.0) using SequenceMatcher."""
    if not a or not b:
        return 0.0
    a_norm = ' '.join(a.lower().split())
    b_norm = ' '.join(b.lower().split())
    return SequenceMatcher(None, a_norm, b_norm).ratio()


def duration_to_seconds(duration_str: str) -> int | None:
    """Convert duration string (M:SS or H:MM:SS) to seconds."""
    if not duration_str:
        return None
    parts = duration_str.split(':')
    try:
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        elif len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except ValueError:
        return None
    return None


def duration_similarity(zing_duration: str, spotify_ms: int) -> float:
    """Compare durations using proportional difference.

    Formula: 1 - |diff| / max(zing_sec, spotify_sec)
    Returns similarity 0.0 to 1.0.
    """
    zing_sec = duration_to_seconds(zing_duration)
    if zing_sec is None or zing_sec == 0:
        return 0.5  # No data, neutral score
    spotify_sec = spotify_ms // 1000
    if spotify_sec == 0:
        return 0.5

    diff = abs(zing_sec - spotify_sec)
    max_duration = max(zing_sec, spotify_sec)
    similarity = 1.0 - (diff / max_duration)

    return max(0.0, similarity)  # Clamp to 0 minimum


def artist_similarity(zing_artists: list[str], spotify_artists: list[dict]) -> float:
    """Compare artist lists, return similarity 0.0 to 1.0."""
    if not zing_artists or not spotify_artists:
        return 0.5  # No data, neutral score

    spotify_names = [a['name'] for a in spotify_artists]

    # Find best match between any ZingMP3 artist and any Spotify artist
    best_match = 0.0
    for zing_artist in zing_artists:
        for spotify_artist in spotify_names:
            sim = string_similarity(zing_artist, spotify_artist)
            best_match = max(best_match, sim)

    return best_match


def calculate_match_score(song: dict, spotify_track: dict) -> float:
    """Calculate comprehensive match score between ZingMP3 song and Spotify track.

    Returns score from 0.0 to 1.0 (higher = better match).
    Uses equal weights for: Title, Artist, Duration (33.3% each)
    """
    title_sim = string_similarity(song['name'], spotify_track['name'])
    artist_sim = artist_similarity(song.get('artist_list', []), spotify_track.get('artists', []))
    duration_sim = duration_similarity(song.get('duration', ''), spotify_track.get('duration_ms', 0))

    # Simple average of 3 fields (equal weights)
    score = (title_sim + artist_sim + duration_sim) / 3

    return score


def search_spotify(sp: spotipy.Spotify, song: dict) -> dict | None:
    """Search for a song on Spotify.

    Args:
        sp: Spotify client
        song: Dict with name, artists, artist_list, duration, url

    Returns dict with all Spotify fields + match_score, or None if not found
    """
    song_name = song['name']
    artists = song.get('artist_list', [])

    # Try different search strategies
    search_queries = []

    # Strategy 1: Full search with song name and first artist
    if artists:
        search_queries.append(f'track:"{song_name}" artist:"{artists[0]}"')

    # Strategy 2: Just song name and artist without quotes
    if artists:
        search_queries.append(f'{song_name} {artists[0]}')

    # Strategy 3: Just song name
    search_queries.append(song_name)

    for query in search_queries:
        try:
            results = sp.search(q=query, type='track', limit=10)  # Get more results to find best match
            tracks = results.get('tracks', {}).get('items', [])

            if tracks:
                # Score each track and pick the best match using comprehensive scoring
                best_track = None
                best_score = -1

                for track in tracks:
                    score = calculate_match_score(song, track)
                    if score > best_score:
                        best_score = score
                        best_track = track

                if best_track:
                    # Convert Spotify duration_ms to M:SS format
                    duration_ms = best_track.get('duration_ms', 0)
                    if duration_ms:
                        minutes = duration_ms // 60000
                        seconds = (duration_ms % 60000) // 1000
                        spotify_duration = f"{minutes}:{seconds:02d}"
                    else:
                        spotify_duration = ''

                    return {
                        'spotify_name': best_track['name'],
                        'spotify_artist': ', '.join([a['name'] for a in best_track['artists']]),
                        'spotify_album': best_track['album'].get('name', ''),
                        'spotify_duration': spotify_duration,
                        'spotify_url': best_track['external_urls'].get('spotify', ''),
                        'popularity': best_track.get('popularity', 0),
                        'track_uri': best_track['uri'],
                        'match_score': best_score
                    }
        except Exception as e:
            print(f"  Search error for '{query}': {e}")
            continue

    return None


def write_excel(results: list[dict], output_file: str) -> None:
    """Write results to Excel file with formatting.

    12 columns: ZingMP3 (5) + Spotify (6) + Match % (1)

    Args:
        results: List of song result dicts
        output_file: Path to output Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ZingMP3 Chart"

    # Headers - ZingMP3 (5 cols) + Spotify (6 cols) + Match (1 col)
    headers = [
        # ZingMP3 data
        'Rank', 'Song Name', 'Artists', 'Duration', 'ZingMP3 Link',
        # Spotify data
        'Song Name (Spotify)', 'Artists (Spotify)', 'Album (Spotify)',
        'Duration (Spotify)', 'Spotify Link', 'Popularity',
        # Match result
        'Match %'
    ]

    # Header styling
    zing_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Blue for ZingMP3
    spotify_fill = PatternFill(start_color='1DB954', end_color='1DB954', fill_type='solid')  # Green for Spotify
    match_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # Orange for Match
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        # Color based on column group
        if col <= 5:
            cell.fill = zing_fill
        elif col <= 11:
            cell.fill = spotify_fill
        else:
            cell.fill = match_fill

    # Data rows
    for row_idx, r in enumerate(results, 2):
        # ZingMP3 columns
        ws.cell(row=row_idx, column=1, value=r['rank'])
        ws.cell(row=row_idx, column=2, value=r['song_name'])
        ws.cell(row=row_idx, column=3, value=r['artists'])
        ws.cell(row=row_idx, column=4, value=r['duration'])
        ws.cell(row=row_idx, column=5, value=r.get('zing_url', ''))

        # Spotify columns
        ws.cell(row=row_idx, column=6, value=r.get('spotify_name', ''))
        ws.cell(row=row_idx, column=7, value=r.get('spotify_artist', ''))
        ws.cell(row=row_idx, column=8, value=r.get('spotify_album', ''))
        ws.cell(row=row_idx, column=9, value=r.get('spotify_duration', ''))
        ws.cell(row=row_idx, column=10, value=r.get('spotify_url', ''))
        ws.cell(row=row_idx, column=11, value=r.get('popularity', 0))

        # Match column (as percentage)
        match_score = r.get('match_score', 0)
        ws.cell(row=row_idx, column=12, value=f"{match_score * 100:.0f}%" if match_score else '')

    # Set column widths
    column_widths = [6, 30, 25, 8, 45, 30, 25, 25, 8, 45, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(output_file)


def main():
    """Main function to crawl chart and find Spotify links."""
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='ZingMP3 Chart Crawler with Spotify Links')
    parser.add_argument('--vpn', action='store_true',
                        help='Fetch live data directly (requires VPN to be connected first)')
    parser.add_argument('--live', action='store_true',
                        help='Fetch live data using Vietnam proxies (unreliable)')
    parser.add_argument('--save-html', action='store_true',
                        help='Save fetched HTML to chart_live.html (only with --vpn or --live)')
    parser.add_argument('--playlist', action='store_true',
                        help='Create/update Spotify playlist with chart songs (requires browser login)')
    parser.add_argument('--playlist-name', type=str, default='ZingMP3 Top 100',
                        help='Name of the Spotify playlist (default: "ZingMP3 Top 100")')
    parser.add_argument('--headless', action='store_true',
                        help='Run in headless/CI mode (uses SPOTIFY_REFRESH_TOKEN env var)')
    parser.add_argument('--chart-url', type=str, default='https://zingmp3.vn/zing-chart',
                        help='URL of the ZingMP3 chart to crawl (default: zing-chart)')
    parser.add_argument('--output-file', type=str, default='zingchart_spotify.xlsx',
                        help='Output Excel filename (default: zingchart_spotify.xlsx)')
    parser.add_argument('--sorted-playlist-name', type=str, default='',
                        help='Create second playlist sorted by Spotify popularity (optional)')
    args = parser.parse_args()

    script_dir = Path(__file__).parent
    html_file = script_dir / 'chart.html'
    output_file = script_dir / args.output_file

    print("=" * 60)
    print("ZingMP3 Chart Crawler with Spotify Links")
    print(f"Chart: {args.chart_url}")
    if args.vpn:
        print("Mode: VPN (direct fetch - VPN must be connected)")
    elif args.live:
        print("Mode: PROXY (fetching via Vietnam proxies)")
    else:
        print("Mode: LOCAL (using saved chart.html)")
    if args.playlist:
        print(f"Playlist: Will create/update '{args.playlist_name}'")
    if args.headless:
        print("Auth: Headless mode (using refresh token)")
    print(f"Output: {output_file}")
    print("=" * 60)

    # Get chart content
    html_content = None
    songs = None  # For weekly charts, we get songs directly from API

    # Check if this is a weekly chart (requires API approach)
    is_weekly = is_weekly_chart_url(args.chart_url)

    if args.vpn:
        # VPN mode - direct fetch (only works for main chart with JSON-LD)
        if is_weekly:
            print("\nWARNING: VPN mode doesn't work for weekly charts (no JSON-LD).")
            print("Please use --live mode for weekly charts.")
        html_content = fetch_zingchart_direct(args.chart_url)
        if not html_content:
            print("\nFailed to fetch via VPN. Make sure VPN is connected!")
            print("Falling back to local chart.html...")

    elif args.live:
        if is_weekly:
            # Weekly chart - use API approach
            songs = fetch_weekly_chart_live(args.chart_url)
            if not songs:
                print("\nFailed to fetch weekly chart via API.")
                import sys
                sys.exit(1)
        else:
            # Main chart - use HTML approach
            html_content = fetch_zingchart_live(args.chart_url)
            if not html_content:
                print("\nFailed to fetch via proxies. Falling back to local chart.html...")

    # Process HTML content or use songs from API
    if songs:
        # Songs already fetched from API (weekly charts)
        print(f"\n[1/3] Using songs from ZingMP3 API...")
    elif html_content:
        # Optionally save HTML
        if args.save_html:
            save_path = script_dir / 'chart_live.html'
            with open(save_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            print(f"  Saved HTML to {save_path}")

        print(f"\n[1/3] Parsing live HTML content...")
        songs = parse_chart_content(html_content)
    else:
        # Use local file - check for chart_live.html first (from proxy fetch), then chart.html
        chart_live_file = script_dir / 'chart_live.html'
        if chart_live_file.exists():
            print(f"\n[1/3] Parsing {chart_live_file} (fetched via proxy)...")
            songs = parse_chart_file(str(chart_live_file))
        elif html_file.exists():
            print(f"\n[1/3] Parsing {html_file}...")
            songs = parse_chart_file(str(html_file))
        else:
            print(f"ERROR: No chart HTML found (checked chart_live.html and chart.html). Cannot continue.")
            import sys
            sys.exit(1)

    print(f"Found {len(songs)} songs")

    # Check Spotify credentials
    if not SPOTIFY_CLIENT_ID or not SPOTIFY_CLIENT_SECRET:
        print("\nERROR: Spotify credentials not found!")
        print("Please set environment variables or create a .env file with:")
        print("  SPOTIFY_CLIENT_ID=your_client_id")
        print("  SPOTIFY_CLIENT_SECRET=your_client_secret")
        return

    # Initialize Spotify
    print("\n[2/3] Connecting to Spotify API...")
    if args.playlist:
        if args.headless:
            print("  (Using refresh token for headless mode)")
            sp = get_spotify_client_headless()
        else:
            print("  (Using user authorization for playlist management)")
            sp = get_spotify_client_with_auth()
    else:
        sp = get_spotify_client()
    print("Connected!")

    # Search Spotify for each song
    print("\n[3/3] Searching Spotify for each song...")
    results = []
    track_uris = []  # Collect track URIs for playlist

    total_songs = len(songs)
    for song in songs:
        pos = song['position']
        name = song['name']

        print(f"  [{pos:3d}/{total_songs}] {name} - {song['artists'][:30]}...", end=" ")

        # Pass full song dict to search_spotify for comprehensive matching
        spotify_result = search_spotify(sp, song)

        if spotify_result:
            match_pct = spotify_result['match_score'] * 100
            print(f"FOUND (match: {match_pct:.0f}%, pop: {spotify_result['popularity']})")
            results.append({
                'rank': pos,
                'song_name': name,
                'artists': song['artists'],
                'duration': song['duration'],
                'zing_url': song.get('url', ''),
                'spotify_name': spotify_result['spotify_name'],
                'spotify_artist': spotify_result['spotify_artist'],
                'spotify_album': spotify_result['spotify_album'],
                'spotify_duration': spotify_result['spotify_duration'],
                'spotify_url': spotify_result['spotify_url'],
                'popularity': spotify_result['popularity'],
                'match_score': spotify_result['match_score'],
                'track_uri': spotify_result['track_uri']
            })
            track_uris.append(spotify_result['track_uri'])
        else:
            print("NOT FOUND")
            results.append({
                'rank': pos,
                'song_name': name,
                'artists': song['artists'],
                'duration': song['duration'],
                'zing_url': song.get('url', ''),
                'spotify_name': '',
                'spotify_artist': '',
                'spotify_album': '',
                'spotify_duration': '',
                'spotify_url': '',
                'popularity': 0,
                'match_score': 0,
                'track_uri': ''
            })

        # Small delay to avoid rate limiting
        time.sleep(0.1)

    # Save to Excel
    print(f"\nSaving results to {output_file}...")
    write_excel(results, str(output_file))

    # Summary
    found_count = sum(1 for r in results if r['spotify_url'])
    print(f"\n{'=' * 60}")
    print(f"Excel saved to: {output_file}")
    print(f"Found on Spotify: {found_count}/{len(results)} songs")

    # Update Spotify playlist if requested
    if args.playlist:
        print(f"\n[4/4] Updating Spotify playlist...")
        playlist_id, is_new = create_or_get_playlist(sp, args.playlist_name, args.chart_url)
        if is_new:
            print(f"  Created new playlist: '{args.playlist_name}'")
        else:
            print(f"  Found existing playlist: '{args.playlist_name}'")

        tracks_added = update_playlist(sp, playlist_id, track_uris)
        print(f"  Added {tracks_added} tracks to playlist")
        playlist_url = f"https://open.spotify.com/playlist/{playlist_id}"
        print(f"  Playlist URL: {playlist_url}")

        # Create second playlist sorted by popularity if specified
        if args.sorted_playlist_name and track_uris:
            print(f"\n[5/5] Creating popularity-sorted playlist...")

            # Sort results by popularity (high to low), keep only found tracks
            sorted_results = sorted(
                [r for r in results if r['spotify_url']],
                key=lambda x: x.get('popularity', 0),
                reverse=True
            )
            sorted_uris = [r['track_uri'] for r in sorted_results]

            sorted_playlist_id, is_new = create_or_get_playlist(
                sp, args.sorted_playlist_name, args.chart_url
            )
            if is_new:
                print(f"  Created new playlist: '{args.sorted_playlist_name}'")
            else:
                print(f"  Found existing playlist: '{args.sorted_playlist_name}'")

            sorted_tracks_added = update_playlist(sp, sorted_playlist_id, sorted_uris)
            print(f"  Added {sorted_tracks_added} tracks (sorted by popularity)")
            sorted_playlist_url = f"https://open.spotify.com/playlist/{sorted_playlist_id}"
            print(f"  Playlist URL: {sorted_playlist_url}")

    print(f"\n{'=' * 60}")
    print("DONE!")
    print(f"{'=' * 60}")

    # Print first 10 as preview
    print("\nPreview (Top 10):")
    print("-" * 100)
    for r in results[:10]:
        status = f"Match: {r['match_score']*100:.0f}%" if r['spotify_url'] else "NOT FOUND"
        print(f"#{r['rank']:2d} | {r['song_name'][:25]:25s} | {r['artists'][:20]:20s} | {status}")


if __name__ == '__main__':
    main()
